import psycopg2
import csv
import xlsxwriter
from openpyxl import load_workbook
from config import host, user, password, db_name
from flask import Flask, render_template, request, send_file
from os import path

try:
    # connect to exist database
    connection = psycopg2.connect(
        host=host,
        user=user,
        password=password,
        database=db_name
    )
    connection.autocommit = True

    cursor = connection.cursor()
    postgreSQL_select_Query = "select * from machine_volume"
    cursor.execute(postgreSQL_select_Query)
    volume_records = cursor.fetchall()

    workbook = xlsxwriter.Workbook('machine_volume.xlsx')
    workbook.close()
    fn = 'machine_volume.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    for row in volume_records:
        ws.append(row)

    wb.save(fn)
    wb.close()


except Exception as _ex:
    print('[INFO] Error while working with PostgreSQL', _ex)
finally:
    if connection:
        connection.close()
        print('[INFO] PostgreSQL connection closed')

# page starting

app = Flask(__name__)


@app.route('/')
def index():

    return render_template('index.html')


@app.route('/success', methods=['POST'])
def success():
    if request.method == 'POST':
        f = request.files['file']
        f.save(f.filename)
        return render_template("success.html", name=f.filename)

@app.route('/dashboard', methods=['POST'])
def dashboard():
    if request.method == 'POST':
        try:
            # connect to exist database
            connection = psycopg2.connect(
                host=host,
                user=user,
                password=password,
                database=db_name
            )
            connection.autocommit = True

            cursor = connection.cursor()
            postgreSQL_truncate_Query = "TRUNCATE table machine_volume RESTART IDENTITY"
            cursor.execute(postgreSQL_truncate_Query)

            with open('volume.csv', newline='') as csvfile:
                reader = csv.DictReader(csvfile, delimiter=';')
                for row in reader:
                    a = row['date']
                    b = row['area']
                    c = row['volume']
                    cursor = connection.cursor()
                    postgreSQL_volume_Query = f"INSERT INTO machine_volume (date, area, volume) VALUES('{a}', {b}, {c});"
                    cursor.execute(postgreSQL_volume_Query)

            with open('machine.csv', newline='') as csvfile:
                reader = csv.DictReader(csvfile, delimiter=';')
                for row in reader:
                    a = row['date']
                    b = row['name']
                    c = row['number']
                    cursor = connection.cursor()
                    postgreSQL_machine_Query = f"Update machine_volume SET name = '{b}', number = '{c}' WHERE date = '{a}';"
                    cursor.execute(postgreSQL_machine_Query)

            cursor = connection.cursor()
            postgreSQL_select_Query = "select * from machine_volume"
            cursor.execute(postgreSQL_select_Query)
            volume_records = cursor.fetchall()


        except Exception as _ex:
            print('[INFO] Error while working with PostgreSQL', _ex)
        finally:
            if connection:
                connection.close()
                print('[INFO] PostgreSQL connection closed')

    return render_template('dashboard.html', volume_records=volume_records)

@app.route('/export', methods=['POST'])
def export():
    if request.method == 'POST':
        try:
            # connect to exist database
            connection = psycopg2.connect(
                host=host,
                user=user,
                password=password,
                database=db_name
            )
            connection.autocommit = True

            cursor = connection.cursor()
            postgreSQL_select_Query = "select * from machine_volume"
            cursor.execute(postgreSQL_select_Query)
            volume_records = cursor.fetchall()

            workbook = xlsxwriter.Workbook('machine_volume.xlsx')
            workbook.close()
            fn = 'machine_volume.xlsx'
            wb = load_workbook(fn)
            ws = wb['Sheet1']
            for row in volume_records:
                ws.append(row)

            wb.save(fn)
            wb.close()


        except Exception as _ex:
            print('[INFO] Error while working with PostgreSQL', _ex)
        finally:
            if connection:
                connection.close()
                print('[INFO] PostgreSQL connection closed')

                return render_template('export.html')

@app.route('/download')
def download():
    return send_file("machine_volume.xlsx", as_attachment=True)

if __name__=='__main__':
    app.run()
